import requests  # For making HTTP requests to Strava API
import urllib3  # To handle SSL warnings for HTTP requests
import pandas as pd  # For handling and processing tabular data
from openpyxl import Workbook, load_workbook  # For creating and editing Excel files
from openpyxl.chart import LineChart, Reference  # For adding charts to Excel
from datetime import datetime  # For working with activity timestamps and formatting dates
import folium # For creating interactive maps
import polyline # For decoding Strava's encoded data route data into GPS coordinates
from openpyxl.utils import get_column_letter # Adjusts column widths within Excel Sheet

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

auth_url = "https://www.strava.com/oauth/token"
activities_url = "https://www.strava.com/api/v3/athlete/activities"

payload = {
    'client_id': "CLIENT_ID",
    'client_secret': 'CLIENT_SECRET',
    'refresh_token': 'REFRESH_TOKEN',
    'grant_type': "refresh_token",
    'f': 'json'
}

print("Requesting Token...\n")
res = requests.post(auth_url, data=payload, verify=False)
access_token = res.json().get('access_token')

if not access_token:
    print("Failed to retrieve access token.")
    exit()

print("Access Token = {}\n".format(access_token))
header = {'Authorization': 'Bearer ' + access_token}

# Pagination setup
request_page_num = 1
all_activities = []

while True:
    param = {'per_page': 200, 'page': request_page_num}
    response = requests.get(activities_url, headers=header, params=param)

    if response.status_code != 200:
        print(f"Error: {response.status_code}, {response.text}")
        break

    my_dataset = response.json()

    if not my_dataset:
        print("No more activities found, stopping fetch.")
        break

    all_activities.extend(my_dataset)
    request_page_num += 1

print(f"Total activities retrieved: {len(all_activities)}")

# Data storage
activity_data = []
kudos_per_month = {}
activities_per_month = {}
total_kudos = 0
total_active_time = 0  # in seconds

for count, activity in enumerate(all_activities, start=1):
    name = activity.get("name", "Unnamed Activity")
    activity_type = activity.get("type", "Unknown")
    moving_time = activity.get("moving_time", 0)  # in seconds
    kudos_count = activity.get("kudos_count", 0)
    start_date = activity.get("start_date", None)

    # Convert start_date to month format
    if start_date:
        activity_date = datetime.strptime(start_date, "%Y-%m-%dT%H:%M:%SZ")
        month_str = activity_date.strftime("%Y-%m")  # Format YYYY-MM
        formatted_date = activity_date.strftime("%Y-%m-%d") # Format for Excel sheet
        # Store kudos count
        kudos_per_month[month_str] = kudos_per_month.get(month_str, 0) + kudos_count
        
        # Count activities per month
        activities_per_month[month_str] = activities_per_month.get(month_str, 0) + 1
    else:
        formatted_time = "N/A"

    # Convert moving_time from seconds to HH:MM:SS format
    hours, remainder = divmod(moving_time, 3600)
    minutes, seconds = divmod(remainder, 60)
    formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"

    # Store data in list
    activity_data.append([count, name, activity_type, formatted_time, kudos_count])

    # Totals
    total_kudos += kudos_count
    total_active_time += moving_time

# Convert total active time to HH:MM:SS format
total_hours, remainder = divmod(total_active_time, 3600)
total_minutes, total_seconds = divmod(remainder, 60)
formatted_total_time = f"{total_hours:02}:{total_minutes:02}:{total_seconds:02}"

# Create DataFrames
df_activities = pd.DataFrame(activity_data, columns=["Activity Number", "Activity Name", "Activity Type", "Active Time", "Kudos"])
df_totals = pd.DataFrame(
    [["Total Kudos", total_kudos], ["Total Active Time", formatted_total_time]],
    columns=["DataType", "Values"]
)
df_kudos_per_month = pd.DataFrame(list(kudos_per_month.items()), columns=["Month", "Kudos"])
df_kudos_per_month.sort_values("Month", inplace=True)
df_activities_per_month = pd.DataFrame(list(activities_per_month.items()), columns=["Month", "Activities"])
df_activities_per_month.sort_values("Month", inplace=True)

# Save to Excel
file_name = "Strava_Activities.xlsx"
with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
    df_activities.to_excel(writer, sheet_name="Activities", index=False)
    df_totals.to_excel(writer, sheet_name="Summary", index=False)
    df_kudos_per_month.to_excel(writer, sheet_name="Kudos Per Month", index=False)
    df_activities_per_month.to_excel(writer, sheet_name="Activities Per Month", index=False)

# Load workbook to add charts
wb = load_workbook(file_name)

# Adjust column widths for "Activities" sheet
ws_activities_auto = wb["Activities"]

for column_cells in ws_activities_auto.columns:
    max_lenth = 0
    column = column_cells[0].column
    column_letter = get_column_letter(column)

    for cell in column_cells:
        try:
            cell_value = str(cell.value)
            if cell_value:
                max_lenth = max(max_lenth, len(cell_value))
        except:
            pass
    adjusted_width = max_lenth + 2
    ws_activities_auto.column_dimensions[column_letter].width = adjusted_width

# Adjust column widths for "Summary" sheet
ws_summary_auto = wb["Summary"]

for column_cells in ws_summary_auto.columns:
    max_lenth = 0
    column = column_cells[0].column
    column_letter = get_column_letter(column)

    for cell in column_cells:
        try:
            cell_value = str(cell.value)
            if cell_value:
                max_lenth = max(max_lenth, len(cell_value))
        except:
            pass
    adjusted_width = max_lenth
    ws_summary_auto.column_dimensions[column_letter].width = adjusted_width


# Kudos Per Month Chart
ws_kudos = wb["Kudos Per Month"]
chart_kudos = LineChart()
chart_kudos.title = "Kudos Over Time"
chart_kudos.x_axis.title = "Month"
chart_kudos.y_axis.title = "Kudos"
data_kudos = Reference(ws_kudos, min_col=2, min_row=1, max_row=ws_kudos.max_row)
categories_kudos = Reference(ws_kudos, min_col=1, min_row=2, max_row=ws_kudos.max_row)
chart_kudos.add_data(data_kudos, titles_from_data=True)
chart_kudos.set_categories(categories_kudos)
ws_kudos.add_chart(chart_kudos, "D2")

# Activities Per Month Chart
ws_activities = wb["Activities Per Month"]
chart_activities = LineChart()
chart_activities.title = "Activities Per Month"
chart_activities.x_axis.title = "Month"
chart_activities.y_axis.title = "Number of Activities"
data_activities = Reference(ws_activities, min_col=2, min_row=1, max_row=ws_activities.max_row)
categories_activities = Reference(ws_activities, min_col=1, min_row=2, max_row=ws_activities.max_row)
chart_activities.add_data(data_activities, titles_from_data=True)
chart_activities.set_categories(categories_activities)
ws_activities.add_chart(chart_activities, "D2")

# Kudos vs Activities Per Month Chart
ws_combined = wb.create_sheet(title="Kudos vc Activities")

# Combine data into one sheet
ws_combined.append(["Month", "Kudos", "Activities"])
for month in sorted(set(kudos_per_month.keys()).union(set(activities_per_month.keys()))):
    ws_combined.append([
        month,
        kudos_per_month.get(month, 0),
        activities_per_month.get(month, 0)
    ])

# Create a line chart
chart_combined = LineChart()
chart_combined.title = "Kudos vs Activities Per Month"
chart_combined.x_axis.title = "Month"
chart_combined.y_axis.title = "Count"

data_combined = Reference(ws_combined, min_col=2, min_row=1, max_col=3, max_row=ws_combined.max_row)
categories_combined = Reference(ws_combined, min_col=1, min_row=2, max_row=ws_combined.max_row)
chart_combined.add_data(data_combined, titles_from_data=True)
chart_combined.set_categories(categories_combined)

ws_combined.add_chart(chart_combined, "E2")

# Save the updated workbook
wb.save(file_name)

print(f"Data and charts successfully written to {file_name}")

# Generate Map
map_center = [0, 0] # Default center, will adjust dynamically
valid_routes = 0
m = folium.Map(location=map_center,
               zoom_start = 12,
               tiles="OpenStreetMap",
               control_scale=True,
               world_copy_jump=False, # Presents wrapping of the map
               no_warp=True, # Disables the infinite scrolling of the map
               max_bounds=[[-90,-180],[90,180]]
)

for activity in all_activities:
    encoded_polyline = activity.get("map", {}).get("summary_polyline", None)

    if encoded_polyline:
        coordinates = polyline.decode(encoded_polyline) #Decode polyline to GPS coordinates
        folium.PolyLine(coordinates, color="red", weight=2.5, opacity=0.8).add_to(m)
        valid_routes += 1

# Adjust the map center if these are valid routes
if valid_routes > 0:
    first_route = polyline.decode(all_activities[0].get("map", {}).get("summary_polyline", ""))
    if first_route:
        map_center = first_route[0]
        m.location = map_center

# Save map to file
map_filename= "strava_activities_map.html"
m.save(map_filename)

print(f"Map saved as {map_filename}")
